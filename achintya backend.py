from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
import random
import os
from datetime import datetime
from apscheduler.schedulers.background import BackgroundScheduler 

app = Flask(__name__)
CORS(app)  # Enable CORS for all routespi

# Data structure for storing usernames and passwords
user_data = {
    'Ryan@123': 'Orion@1',
    'Subbaiah@123': 'Orion@2',
    'Allen@123': 'Orion@3',
    'Achintya@123': 'Orion@4',
    'Prajwal@123': 'Orion@5',
    'Tanisha@123': 'Orion@6'
}

# Global variable for file name
file_name = 'inspection_data.xlsx'

@app.route('/authorization', methods=['POST'])
def login():
    data = request.get_json()

    # Extract username and password from the request
    username = data.get('username')
    password = data.get('password')

    # Authenticate the user
    if username in user_data and user_data[username] == password:
        # If authentication is successful, set control variable to 1
        response_data = {'control_variable': 1, 'message': 'Authentication successful'}
    else:
        # If authentication fails, set control variable to 0
        response_data = {'control_variable': 0, 'message': 'Authentication failed'}

    # Return the response as JSON
    return jsonify(response_data)

# Data structure for storing inspection data
inspection_data = []

# Create a scheduler
scheduler = BackgroundScheduler()
scheduler.start()

# Global variable for file name
file_name = 'inspection_data.xlsx'

# Endpoint to receive inspection data and save it to an Excel file
@app.route('/saveToExcel', methods=['POST'])
def save_to_excel():
    global file_name, inspection_data

    data1 = request.get_json()

    # Append new data to the list
    inspection_data.extend(data1)

    # Create a new Excel workbook or load an existing one
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active

        # Write header to the sheet
        ws.append(['Pulley ID', 'Oil Contamination', 'Foreign Contamination', 'Cracks', 'Lagging', 'Date'])
    else:
        # Load the existing Excel sheet
        wb = load_workbook(file_name)
        ws = wb.active

    # Write data to the sheet
    for row in inspection_data:
        ws.append([row['pulleyID'], row['oilContamination'], row['foreignresidue'], row['cracks'], row['lagging'], row['date']])

    # Save the workbook to a file
    wb.save(file_name)

    # Clear inspection_data to avoid duplicates
    inspection_data.clear()

    return jsonify({'success': True, 'message': 'Data saved to Excel file.'})

# Function to generate and save random values for temperature and humidity
def generate_random_values(file_name):
    global inspection_data

    # Load the existing Excel sheet
    wb = load_workbook(file_name)
    ws = wb.active

    # Generate and save random values for each row
    for row in inspection_data:
        pulley_id = row['pulleyID']
        temperature = round(random.uniform(20, 30), 2)
        humidity = round(random.uniform(30, 70), 2)
        #add other variables
        # Append new values to the row
        ws.append([pulley_id, row['oilContamination'], row['foreignresidue'], row['cracks'], row['lagging'], row['date'], temperature, humidity])

    # Save the workbook to a file
    wb.save(file_name)

# Schedule the function to run every hour
scheduler.add_job(generate_random_values, 'interval', args=[file_name], hours=1)

if __name__ == '__main__':
    app.run(debug=True)



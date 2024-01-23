from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from datetime import datetime
from apscheduler.schedulers.background import BackgroundScheduler
import random
import os

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

# Data structure for storing usernames and passwords
user_data = {
    'Ryan@123': 'Orion@1',
    'Subbaiah@123': 'Orion@2',
    'Allen@123': 'Orion@3',
    'Achintya@123': 'Orion@4',
    'Prajwal@123': 'Orion@5',
    'Tanisha@123': 'Orion@6'
}

@app.route('/authorization', methods=['POST'])
def login():
    data = request.get_json()

    # Extract username and password from the request
    username = data.get('username')
    password = data.get('password')

    # Authenticate the user
    if username in user_data and user_data[username] == password:
        # If authentication is successful, set control variable to 1
        response_data = {'control_variable': 1, 'message': 'Authentication successful'}
    else:
        # If authentication fails, set control variable to 0
        response_data = {'control_variable': 0, 'message': 'Authentication failed'}

    # Return the response as JSON
    return jsonify(response_data)


# Data structure for storing inspection data
inspection_data = []

# Global variable for file name
file_name = 'inspection_data.xlsx'


# Create a scheduler
scheduler = BackgroundScheduler()
scheduler.start()


# Function to generate and save random values for temperature and humidity
def generate_random_values(file_name):
    global inspection_data

    # Load the existing Excel sheet
    wb = load_workbook(file_name)
    ws = wb.active

    # Generate and save random values for each row
    for row in inspection_data:
        pulley_id = row['pulleyID']
        temperature = round(random.uniform(20, 30), 2)
        humidity = round(random.uniform(30, 70), 2)
        # Add other variables
        # Append new values to the row
        ws.append([pulley_id, row['oilContamination'], row['foreignresidue'], row['cracks'], row['lagging'],
                   row['date'], temperature, humidity])

    # Save the workbook to a file
    wb.save(file_name)

# Schedule the function to run every hour
scheduler.add_job(generate_random_values, 'interval', args=[file_name], hours=1)

# ...

# Endpoint to receive inspection data and save it to an Excel file
@app.route('/saveToExcel', methods=['POST'])
def save_to_excel():
    global file_name, inspection_data

    try:
        print("Endpoint /saveToExcel reached")
        data1 = request.get_json()

        # Convert the incoming data to a list of dictionaries
        inspection_data.extend(data1)

        # Specify an absolute file path
        file_path = os.path.join(os.path.abspath(os.path.dirname(__file__)), file_name)

        # Create a new Excel workbook or load an existing one
        if not os.path.exists(file_path):
            wb = Workbook()
            ws = wb.active

            # Write header to the sheet
            ws.append(['Pulley ID', 'Oil Contamination', 'Foreign Contamination', 'Cracks', 'Lagging', 'Date'])
        else:
            # Load the existing Excel sheet
            wb = load_workbook(file_path)
            ws = wb.active

        # Write data to the sheet
        for row in inspection_data:
            # Ensure row is a dictionary
            if isinstance(row, dict):
                ws.append([row.get('pulleyID', ''), row.get('oilContamination', ''), row.get('foreignresidue', ''),
                           row.get('cracks', ''), row.get('lagging', ''), row.get('date', '')])

        # Save the workbook to a file
        wb.save(file_path)

        # Clear inspection_data to avoid duplicates
        inspection_data.clear()

        print("Received data:", data1)
        return jsonify({'success': True, 'message': 'Data saved to Excel file.'})
    except Exception as e:
        print("Error saving data:", str(e))
        return jsonify({'success': False, 'message': 'Error saving data. Please try again.'})

# ...
from twilio.rest import Client
account_sid = 'AC56f010049aa18eb118b7824b588d8172'
auth_token = '8b28bedfab535da73562e1fb6adfe8df'
twilio_phone_number = '+12058528442'
client = Client(account_sid, auth_token)

@app.route('/send-sms', methods=['POST'])
def send_sms():
    data = request.get_json()

    to = data.get('to')
    body = data.get('body')

    if to and body:
        try:
            message = client.messages.create(
                to=to,
                from_=twilio_phone_number,
                body=body
            )
            print(f"SMS sent successfully. SID: {message.sid}")
            return jsonify(success=True, message="SMS sent successfully")
        except Exception as e:
            print(f"Error sending SMS: {str(e)}")
            return jsonify(success=False, message="Error sending SMS"), 500
    else:
        return jsonify(success=False, message="Missing 'to' or 'body' parameter"), 400
from flask import Flask, request, jsonify
from joblib import load
import random
model = load('random_forest_model1.joblib')

def generate_random_sensor_values():
    return [
        random.choice(['1', '2', '3', '4', '5']),
        random.choice([0, 1, 2, 3, 4, 5, 6]),
        random.uniform(2000, 20000),
        random.choice([1900, 1800, 1700, 1500, 1100]),
        random.choice(['1', '2', '3', '4']),
        random.uniform(30, 50),
        random.uniform(30, 70),
        random.uniform(1, 5),
        random.choice(['1', '2', '3', '4']),
        random.choice(['1', '2', '3', '4']),
        random.choice([0.5, 0.7, 0.8, 0.4]),
        random.choice([3000, 1500, 600, 5000, 8000]),
        random.uniform(1, 40),
        random.choice(['1', '2', '3', '4']),
        random.choice(['1', '2', '3', '4']),
        random.choice([8, 6, 10, 12]),
    ]

@app.route('/your-backend-endpoint', methods=['POST'])
def receive_data():
    try:
        # Get the data from the request
        data = request.json

        # Check if the data is 1
        if data == 1:
            print('Received data:', data)
            
            # Generate random sensor values
            sensor_data = generate_random_sensor_values()
            
            # Load the model and predict
            prediction = model.predict([sensor_data])

            # Determine the response based on the prediction
            if prediction == 1:
                response = 'Alert'
            else:
                response = 'No Alert'

            # Print the prediction
            print("Model Prediction:", response)
            pid=random.choice(['DV1','DV2','DV3','DV4','DV5','SV1','SV2','SV3'])
            # Return the prediction to the frontend
            return jsonify({ 'prediction': response, 'pulleyID': pid})
        else:
            return jsonify({'error': 'Invalid data'})

    except Exception as e:
        print('Error:', e)
        return jsonify({'error': 'Internal server error'})

if __name__ == '__main__':
    app.run(debug=True, host='192.168.84.116', port=5000)

